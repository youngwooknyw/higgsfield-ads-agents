from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "DXRS75063-WHS PDA"

thin = Side(border_style="thin", color="C0C0C0")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
hdr_fill = PatternFill("solid", fgColor="1F3A5F")
hdr_font = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=10)
sub_font = Font(name="맑은 고딕", bold=True, size=10)
cell_font = Font(name="맑은 고딕", size=10)
prompt_font = Font(name="Consolas", size=9)
wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

info = [
    ("상품", "DXRS75063-WHS (유니 프레시벤트 세미오버핏 스몰로고 반팔티셔츠 / WHITE)"),
    ("카테고리", "RS (반팔티셔츠) / 26SS"),
    ("POINT", "스몰와펜으로 포인트를 준 세미오버핏 베이직 반팔 티셔츠"),
    ("FABRIC", "FRESHVENT 접촉냉감 + UV차단 (야외활동 적합)"),
    ("FIT", "세미오버핏 - 편안한 착용감과 활동성 / 여유있는 핏"),
]
for i, (k, v) in enumerate(info, start=1):
    ws.cell(row=i, column=1, value=k).font = sub_font
    ws.cell(row=i, column=2, value=v).font = cell_font

ws.cell(row=7, column=1, value="데이터 소스").font = sub_font
ws.cell(row=7, column=2, value="Crema API 54건(WHS 11건) | DX 26SS 스타일가이드 | 무신사 26SS 트렌드(반소매 검색량 6배↑, 얼리서머, 쿨링 원단)").font = cell_font
ws.cell(row=8, column=1, value="리뷰 빈도 상위").font = sub_font
ws.cell(row=8, column=2, value="시원함/여름(18) > 소재/품질(15) > 핏/사이즈(12) > 착용감/편안함(10) > 디자인/색상(10) > 배송/CS(6) > 재구매(3) > 가성비(2)").font = cell_font
ws.cell(row=9, column=1, value="WHS 특이 시그널").font = sub_font
ws.cell(row=9, column=2, value="화이트 → 청량감↑·이너룩 활용↑ / '부모님 선물' 페르소나 신호 / '무신사 구매자' 외부 채널 신호").font = cell_font

for r in range(1, 10):
    for c in range(1, 3):
        ws.cell(row=r, column=c).border = border

HEADER_ROW = 12
headers = [
    ("#", 4),
    ("Persona", 20),
    ("Persona 상세", 26),
    ("Desire 계층", 11),
    ("Desire 내용", 22),
    ("Awareness", 14),
    ("핵심 카피 후킹", 32),
    ("비주얼 디렉션", 32),
    ("힉스필드 프롬프트 (EN)", 80),
    ("브랜드 적합성\n(O/X/△)", 12),
    ("우선순위\n(1~8)", 9),
    ("진행 여부", 10),
    ("코멘트", 24),
]
for ci, (label, w) in enumerate(headers, start=1):
    c = ws.cell(row=HEADER_ROW, column=ci, value=label)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = center
    c.border = border
    ws.column_dimensions[get_column_letter(ci)].width = w

SAFETY = "Photorealistic, premium e-commerce campaign aesthetic, subtle film grain, 35mm look. No people, no extra logos, no text in image."

prompts = [
    "Editorial product still-life for Discovery Expedition Korea, FRESHVENT cooling-touch story for the morning commute. Hero: the same white short-sleeve t-shirt as the reference image, semi-over fit, small embroidered chest patch (do not render readable logo text). The white tee laid flat and slightly rumpled on a brushed cool-grey metal surface evoking a subway handrail texture. Visual metaphors of coolness: tiny condensation droplets dotted across the metal, a thin trail of cold morning mist drifting low across the bottom of the frame, delicate frost-crystal motifs barely visible on the fabric surface. Lighting: cold blue-white morning daylight from upper-left, sharp shadow edges to imply freshness, subtle warm highlight on the fabric weave. Background: softly out-of-focus cool-blue gradient suggesting blurred train interior. Composition: 3/4 top-down, tee centered slightly right, generous clean negative space at top for one-line Korean ad copy. " + SAFETY,

    "Editorial product still-life for Discovery Expedition Korea, parents-and-children summer gifting narrative. Hero: two of the same white short-sleeve t-shirts as the reference image (one slightly larger, one slightly smaller, both neatly folded), semi-over fit, small embroidered chest patches visible (do not render readable logo text). The two folded tees arranged side by side inside an opened warm-beige craft gift box on a natural-wood table. Gifting props around: folded brown kraft paper, a sprig of dried wheat, a small blank handwritten card, a soft cream linen ribbon untied beside the box. Lighting: warm soft morning light from window-right, gentle long shadows, warm color temperature suggesting care and family warmth. Background: out-of-focus cream wall and a hint of indoor plant. Composition: top-down 3/4 angle, gift box dead-center, generous clean negative space at lower-third for Korean ad copy. " + SAFETY,

    "Editorial product still-life for Discovery Expedition Korea, minimalism narrative for the modern urban wardrobe. Hero: the same white short-sleeve t-shirt as the reference image, semi-over fit, small embroidered chest patch (do not render readable logo text), hung on a thin natural-wood hanger. Beside it: a pair of clean indigo selvedge denim hung on an identical hanger. Both hangers on a slim matte-black wall rail mounted on a smooth pure-white plaster wall. Nothing else in the frame, extreme negative space. Lighting: soft diffused daylight from off-frame left, subtle shadow cast by the tee, low contrast, near-monochrome warm-grey palette. Composition: head-on, slightly above eye-level, two garments centered together at the lower third of the frame, with massive clean negative space above for Korean ad copy. " + SAFETY,

    "Editorial product still-life for Discovery Expedition Korea, dual-function (UV protection + cooling-touch) story for active outdoor runners along the Han River. Hero: the same white short-sleeve t-shirt as the reference image, semi-over fit, small embroidered chest patch (do not render readable logo text). The white tee laid casually on smooth river-walk concrete after a run. Beside it: a stainless steel water bottle with fresh condensation, a pair of running gloves, a folded thin breathable cap, a small smartwatch face-up. Fabric shows soft natural folds suggesting recent wear. Lighting: strong midday sun from directly above, crisp hard shadow with bright highlights, hot sunlight contrasted by cool grey concrete. A small soft halo of bright light on the white fabric to suggest UV protection. Background: out-of-focus glittering river surface and concrete walkway. Composition: top-down 3/4, tee left-center, generous clean negative space at right for two-line Korean ad copy. " + SAFETY,

    "Editorial product still-life for Discovery Expedition Korea, weekend family-outing narrative emphasizing wash-and-wear durability for a dad daily tee. Hero: the same white short-sleeve t-shirt as the reference image, semi-over fit, small embroidered chest patch (do not render readable logo text). The white tee laid spread out on a folded picnic blanket of warm beige and rust-red stripes on a sunlit grass field. Surrounding props: a worn-in leather camping bag half-open, a small enamel mug, a folded brown linen towel, a half-eaten apple, a paperback book. Quiet weekend-after-the-hike vibe. Soft dappled tree shadows fall across the blanket and the tee. Lighting: warm late-morning sun filtering through leaves, golden hour warmth, soft natural shadows. Background: out-of-focus green meadow with a hint of distant mountain ridge. Composition: top-down 3/4, tee center, generous clean negative space at top for Korean ad copy. " + SAFETY,

    "Editorial product still-life for Discovery Expedition Korea, value-for-money comparison narrative for the marketplace-platform shopper. Hero: the same white short-sleeve t-shirt as the reference image, semi-over fit, small embroidered chest patch (do not render readable logo text). The white tee laid flat and crisp on a clean light-grey paper surface, perfectly centered. Surrounding the tee: clean empty white rectangular panels arranged symmetrically like a feature comparison sheet (panels are completely blank, leave empty for post-production text overlay), and a single small bright marigold-yellow accent dot in the corner of the frame for color punch. Lighting: bright even studio daylight, no harsh shadow, product-photo crisp. Background: pure light-grey paper texture. Composition: top-down 90-degree overhead, tee dead-center, blank panels arranged symmetrically left and right with generous negative space below for Korean price and feature copy. Photorealistic, clean e-commerce comparison aesthetic, sharp. No people, no extra logos, no text in image.",

    "Editorial product still-life for Discovery Expedition Korea, two-way styling narrative (inner-layer plus solo-wear) for the design-forward urbanite. Hero: two arrangements of the same white short-sleeve t-shirt as the reference image side by side in one frame, semi-over fit, small embroidered chest patch visible (do not render readable logo text). LEFT half: the white tee styled as an inner layer, laid flat with a soft sand-beige oversized linen overshirt loosely draped over its shoulders. RIGHT half: the same white tee styled standalone, laid flat with a folded charcoal-grey wide-leg trouser placed neatly beside it. Background: smooth deep-charcoal slate stone surface, monochrome low-saturation palette with the white tee as the bright focal point in both halves. Lighting: even soft directional daylight from above, gallery-style minimal shadow. Composition: top-down 90-degree overhead, two arrangements separated by a subtle vertical line of negative space down the center. Generous clean negative space at top for Korean ad copy. " + SAFETY,

    "Editorial product still-life for Discovery Expedition Korea, science-lab inspired teaser for FRESHVENT cooling-touch. Provocative hook for an audience that has never tried functional cooling apparel. Hero: the same white short-sleeve t-shirt as the reference image, semi-over fit, small embroidered chest patch (do not render readable logo text). The white tee laid flat on a pure matte-black surface, with a thin sweeping beam of cool blue laboratory light passing diagonally across the fabric (suggesting a thermal scan), small floating frost-particle motes rising from the surface, a tiny pool of clear water beading on the matte-black just beside the tee suggesting moisture absorption. No physical text or numbers visible on the image, leave blank for post-production thermal-graphic overlay. Lighting: dramatic side-lighting from upper-right, cool 5500K blue-white tone, very high contrast black-and-white palette with the white tee as the only bright element. Background: pure matte black with a hint of soft cold light fog drifting low. Composition: top-down 3/4 angle, tee center, vast clean negative black space at left and bottom for Korean teaser copy. Photorealistic, science-lab campaign aesthetic, sharp edges. No people, no extra logos, no text in image.",
]

pda = [
    (1, "30대 직장인 남성", "도시 출근, 만원 지하철 통근, 셔츠 안에 이너로도 활용", "기능", "더위 해방 (FRESHVENT 접촉냉감)", "Problem-aware", "출근길 만원 지하철, 등판이 다르다", "화이트 티 단독, 미스트+얼음결정 오버레이, 차가운 블루 톤"),
    (2, "40대 자녀 (부모님 선물러)", "부모 환갑·여름맞이 효도 선물, 브랜드 신뢰 우선", "사회", "가족에게 안전한 브랜드 선물", "Most-aware", "올여름, 부모님께도 디스커버리 한 장", "두 사이즈 화이트 티 + 선물 박스, 따뜻한 베이지 톤, 손글씨 카드"),
    (3, "20대 후반 미니멀리스트", "서울 1인가구, 옷장 최소화, OOTD SNS 활동", "정체성", "깔끔한 자기표현 (단순함의 미학)", "Solution-aware", "옷장은 단순해질수록 자유로워진다", "미니멀 룸, 옷걸이에 화이트 티 + 데님, 흑백 톤, 여백 강조"),
    (4, "액티브 입문자", "주말 한강 러닝·하이킹 입문, 30대 초반", "기능", "야외 신뢰할 듀얼 기능성 (UV+쿨링)", "Solution-aware", "한낮의 한강, 자외선부터 땀까지 한 번에", "한강변 러닝 직후, 강한 햇살, UV/FRESHVENT 인포 그래픽"),
    (5, "30대 가장 (주말 가족 외출)", "두 자녀 부모, 캠핑·피크닉 등 가족 단위 활동", "편의", "매일 빨아도 변형 없는 데일리", "Problem-aware", "이번 주말, 또 같은 티? — 매일 빨아도 변형 없는 데일리", "가족 캠핑·피크닉 라이프스타일 씬, 자연광"),
    (6, "무신사 가성비 쇼퍼", "20~30대, 무신사 채널 메인, 가성비 비교 쇼퍼", "경제", "이 가격에 냉감 기능 (가성비)", "Most-aware", "이 가격에 냉감? 디스커버리니까", "인포그래픽, 가격×기능 비교 표, 옐로/블랙 강조"),
    (7, "디자이너/크리에이티브 직군", "셔츠 페어링 미니멀리스트, 이너룩~단독 양용", "사회/정체성", "트렌디 OOTD (한 벌 두 룩)", "Product-aware", "셔츠 이너로 한 장, 단독으로 한 장 — 한 벌 두 룩", "분할 컷 이너룩 vs 단독, 도시 무드, 모노톤"),
    (8, "디스커버리 미경험 30대 남성", "타 아웃도어 브랜드 사용 중, 디스커버리 인지 낮음", "기능", "더위 해방 (호기심 자극)", "Unaware", "에어컨 옷, 입어보셨어요?", "사이언스 톤, 검정 배경 + 데이터 시각화"),
]

start_row = HEADER_ROW + 1
for ri, row in enumerate(pda):
    rr = start_row + ri
    for ci, value in enumerate(row, start=1):
        c = ws.cell(row=rr, column=ci, value=value)
        c.font = cell_font
        c.alignment = wrap
        c.border = border
        if ci == 1:
            c.alignment = center
            c.font = sub_font
    pc = ws.cell(row=rr, column=9, value=prompts[ri])
    pc.font = prompt_font
    pc.alignment = wrap
    pc.border = border
    for ci in range(10, 14):
        c = ws.cell(row=rr, column=ci, value="")
        c.font = cell_font
        c.alignment = wrap
        c.border = border

for ri in range(len(pda)):
    ws.row_dimensions[start_row + ri].height = 240

dv1 = DataValidation(type="list", formula1='"O,X,△"', allow_blank=True)
dv1.add(f'J{start_row}:J{start_row+len(pda)-1}')
ws.add_data_validation(dv1)
dv2 = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8"', allow_blank=True)
dv2.add(f'K{start_row}:K{start_row+len(pda)-1}')
ws.add_data_validation(dv2)
dv3 = DataValidation(type="list", formula1='"진행,보류,제외"', allow_blank=True)
dv3.add(f'L{start_row}:L{start_row+len(pda)-1}')
ws.add_data_validation(dv3)

ws.freeze_panes = f'A{HEADER_ROW+1}'

out = r'C:/Users/AD0899/higgsfield ads/DXRS75063-WHS_PDA_lineup.xlsx'
wb.save(out)
print(f'Saved: {out}')
